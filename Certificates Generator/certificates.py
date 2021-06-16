import subprocess
# The subprocess module present in Python is used to run new applications or programs through Python code by creating new processes.
import sys
# The sys module provides functions and variables used to manipulate different parts of the Python runtime environment. 
import os
# The OS module in Python provides functions for interacting with the operating system. 

def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])
# The officially recommended way to install packages from a script is by calling pip's command-line interface via a subprocess.

install('img2pdf')
# install img2pdf python package for Lossless conversion of raster images to PDF
install('openpyxl')
# install openpyxl pyhton library to read/write Excel 2010 xlsx/xlsm/xltx/xltm files
install('Pillow')
# install pillow that is a fork of Python Imaging Library (PIL), which adds support for opening, manipulating, and saving images.

from img2pdf import convert, AlphaChannelError
# Image can be converted into pdf bytes using img2pdf.convert() function 
# An alpha channel is to process overlaying a foreground image with transparency over a background image.
from PIL import Image, ImageDraw, ImageFont
#PIL is the Python Imaging Library which provides the python interpreter with image editing capabilities.
# The module also provides a number of factory functions, including functions to load images from files, and to create new images.
# The ImageDraw module provide simple 2D graphics for Image objects. 
# The ImageFont module defines a class that store bitmap fonts, and are used with the PIL. 
from tkinter.filedialog import askopenfile
# The tkinter.filedialog module provides classes and factory functions for creating file/directory selection windows.
# askopenfile is a file opener function to open and read any text based files ike .txt or .csv files.
from openpyxl import load_workbook
# openpyxl is a Python library to read/write Excel (with extension xlsx/xlsm/xltx/xltm) files.
# load workbook function to open already created workbook or file on your disk for some operation.

file = askopenfile(title='Select the Workbook', mode='r', filetypes=[
                   ('Microsoft Excel', '.xlsx .xlsx .xlsm .xltx .xltm')])
# This function will be used to open file in read mode and only Excel files will be opened
if file is not None:
    dirpath = os.path.dirname(file.name)
else:
    sys.exit()
# os.path.dirname() method is used to get the directory name from the specified path.
# sys.exit() to exit from python 
filepath = file.name

allCertPath = dirpath+"/All_Certificates"
# allCerpath holds the path of All_Certificates folder to be created
allCertImgPath = allCertPath+"/Images"
# allCertImgPath holds the path of Images subfolder to be created
allCertPdfPath = allCertPath+"/PDFs"
# allCertImgPath holds the path of PDFs subfolder to be created


try:
    os.mkdir(allCertPath)
    os.mkdir(allCertImgPath)
    os.mkdir(allCertPdfPath)
# os.mkdir() method is used to create a directory named with the specified path.This method raise FileExistsError if the directory to be created already exists.
# All_Certificates folder and Images and PDFs subfolders are created in the specified path 
except:
    {}

wb = load_workbook(filepath,data_only=True)
# The workbook is opened and accessed via wb object
for ws in wb:
# ws object can access all worksheets of the workbook
# for each worksheet in the workbook
    for r in range(3, 6):
    # r object can access all rows mentioned in the range from the worksheet 
    # for all rows numbered from 3 to 200
        cell = ws.cell(row=r, column=1)
    # Cell objects also have a row, column,and coordinate attributes that provide location information for the cell.
    # for every row cell object stores data of specified by respective row and column number 1
        if cell.value is None:
            break

        name = str(ws.cell(row=r, column=2).value).strip().title()
        # name object stores the name of the member
        acm_id = str(ws.cell(row=r, column=1).value).strip().upper()
        # acm_id object stores the id(the unique number) of the member
        # strip() is an inbuilt function that returns a copy of the string with both leading and trailing characters removed.
        # The title() function is used to convert the first charalcter in each word to Uppercase and remaining characters to Lowercase in the string.
        # The upper() methods converts all lowercase characters of the string to uppercase.
        eachmemberIMGpath = allCertImgPath+'/'+acm_id+'.png'
        # eachmemberIMG object holds the path for acm_id.png file
        eachmemberPDFpath = allCertPdfPath+'/'+acm_id+'.pdf'
        # eachmemberPDF object holds the path for acm_id.pdf file
        certificate = Image.open('certificates.png')
        # certificate objects holds the sample template certificates.png
        # image is opened in certificate object
        draw = ImageDraw.Draw(certificate)
        # draw object holds the image context that need to be modified
        name_font = ImageFont.truetype('Lora-Bold.ttf', 75)
        # ImageFont.truetype() loads a font object from the given file, and creates a font object for a font of the given size.
        # name_font object holds specified font style and size
        w, h = draw.textsize(name, name_font)
        left = (certificate.width - w) / 2
        top = 550
        # left and top objects specify the left and top coordinates of the sample template where the name of the member is to be printed
        draw.text((left, top), name, fill=(75, 75, 75, 255), font=name_font)
        # data specified in name object is written on the image at specified coordinates
        certificate = certificate.convert('RGB')
        # Image.convert('RGB') just converts each pixel to the triple 8-bit value,it basically changes the mode of how image is represented and stored.
        certificate.save(eachmemberIMGpath)
        # The changes done to the sample template are saved and the respective image is stored in path specified by eachmemberIMGpath object
        certificate = Image.open(eachmemberIMGpath)
        # image in eachmemberIMGpath is opened in certificate object
        pdf_bytes = convert(certificate.filename)
        # convert the image file into cunks using covert method
        f = open(eachmemberPDFpath, "wb")
        # open the pdf file
        f.write(pdf_bytes)
        # write the pdf file with chunks
        certificate.close()
        # close the image file
        f.close()
        # The respective image file is converted to pdf file and stored in eachmemberPDFpath
        print(r, name)
        # row number and name of the member are printed on successful generation of certificate in .png and .pdf formats
        # it continues with the next row
    # it continues with the next sheet in wb
wb.save(file.name)
# The workbook is saved and closed
print("Done.")
# After the successful execution of the program a Done. statement is printed in the logs...!!!